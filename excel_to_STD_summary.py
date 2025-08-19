"""
Create pointwise STD/CoV summary files from per-trial Excel workbooks.

Each input file is expected to be named like:
    <trial>_..._STD_CoV_<force>[_...].xlsx
Examples:
    007_trial_STD_CoV_FgR_sum.xlsx
    003_trial_STD_CoV_FgR_sum.xlsx

Within each workbook there are two sheets (one per side).
Each sheet contains pointwise columns (Mean, STD, CoV) and the last three
rows summarize the data with labels like 'min', 'max', 'mean'.
We extract from each sheet:
    - meanSTD  := value in row 'mean', column containing STD
    - meanCOV  := value in row 'mean', column containing CoV
    - maxSTD   := value in row 'max',  column containing STD
    - maxCoV   := value in row 'max',  column containing CoV

For each distinct <force> found in the folder, we create one summary file:
    pointwise_STD_<force>_summary.xlsx
containing two sheets (one per side). Each sheet has the rows for trials with
columns: Trial, meanSTD, meanCOV, maxSTD, maxCoV.
"""

from __future__ import annotations

import os
import re
from pathlib import Path
from typing import Dict, List, Tuple, Optional, Iterable
import config

import pandas as pd

def _iter_excel_files(root: Path, recursive: bool = False) -> Iterable[Path]:
    """Yield .xlsx/.xls files from a folder (optionally recursively)."""
    if recursive:
        for p in root.rglob("*"):
            if p.is_file() and p.suffix.lower() in {".xlsx", ".xls"}:
                yield p
    else:
        for p in root.iterdir():
            if p.is_file() and p.suffix.lower() in {".xlsx", ".xls"}:
                yield p

# --- Configuration (edit if needed) -----------------------------------------
# Column-name patterns for the statistics columns (case-insensitive matching)
STD_COL_PATTERNS = (
    r"^std$",
    r"standard\s*deviation",
)
COV_COL_PATTERNS = (
    r"^cov$",
    r"^cov\b.*",  # e.g., 'CoV [%]'
    r"co[v|efficient]\s*of\s*variation",
)
# Summary row labels (case-insensitive)
MEAN_ROW_PATTERNS = (r"^mean$",)
MAX_ROW_PATTERNS = (r"^max$",)
# ---------------------------------------------------------------------------


def _find_first_matching(name_list: List[str], patterns: Tuple[str, ...]) -> str | None:
    """Return the first name from name_list that matches any regex in patterns (case-insensitive)."""
    for pat in patterns:
        rx = re.compile(pat, flags=re.IGNORECASE)
        for name in name_list:
            if isinstance(name, str) and rx.search(name.strip()):
                return name
    return None


def _extract_force_from_filename(fname: str) -> str | None:
    """Extract the force token that follows 'CoV_' in the filename.

    Example: '007_trial_STD_CoV_FgR_sum.xlsx' -> 'FgR_sum'
    """
    m = re.search(r"STD_CoV_([^/]+?)(?=\.xlsx$)", fname)
    return m.group(1) if m else None


def _extract_trial_from_filename(fname: str) -> str:
    """Derive a compact trial identifier from the filename.

    Preference order:
    1) Leading number sequence (e.g., '007')
    2) Token before first underscore
    3) Whole stem
    """
    stem = Path(fname).stem
    m = re.match(r"(\d+)", stem)
    if m:
        num = m.group(1)
        if num in config.file_acronyms_map:
            return f"{num}_{config.file_acronyms_map[num]}"
        return num
    parts = stem.split("_")
    return parts[0] if parts else stem


def _pick_summary_values(df: pd.DataFrame) -> Tuple[float, float, float, float]:
    """From a single sheet DataFrame, pick meanSTD, meanCOV, maxSTD, maxCoV.

    The sheet is expected to contain summary rows labeled like 'min', 'max', 'mean'.
    We try to locate rows by label (in index or first column). If no labels are
    found, we assume the *last three rows* are [min, max, mean] in that order.
    """
    # Ensure we have a copy to avoid SettingWithCopy warnings
    df = df.copy()

    # If the index looks numeric and we have a labeled first column, use that as index
    if df.index.dtype != object:
        first_col = df.columns[0]
        if df[first_col].dtype == object:
            df.set_index(first_col, inplace=True)

    # Try to normalize index to string labels for matching
    idx_labels = [str(x).strip() if pd.notna(x) else "" for x in df.index]
    df.index = idx_labels

    # Identify the key columns for STD and COV
    cols = [str(c).strip() for c in df.columns]
    std_col = _find_first_matching(cols, STD_COL_PATTERNS)
    cov_col = _find_first_matching(cols, COV_COL_PATTERNS)
    if std_col is None or cov_col is None:
        raise ValueError("Could not locate STD/COV columns in sheet. Columns found: " + ", ".join(cols))

    # New rule: summary rows are the last two rows
    # - second-to-last row = max
    # - last row = mean
    # We first drop completely empty rows to guard against trailing blanks.
    df_clean = df.copy()
    df_clean = df_clean.dropna(how="all")
    if df_clean.shape[0] < 2:
        raise ValueError("Not enough rows to read summary (need at least two rows for max/mean at the end).")
    tail2 = df_clean.tail(2)
    # Ensure stable order: second-to-last first, then last
    second_last_idx, last_idx = list(tail2.index)[0], list(tail2.index)[1]

    # Convert to numeric safely
    maxSTD_val = pd.to_numeric(tail2.loc[second_last_idx, std_col], errors="coerce")
    meanSTD_val = pd.to_numeric(tail2.loc[last_idx, std_col], errors="coerce")
    maxCOV_val = pd.to_numeric(tail2.loc[second_last_idx, cov_col], errors="coerce")
    meanCOV_val = pd.to_numeric(tail2.loc[last_idx, cov_col], errors="coerce")

    if pd.isna(meanSTD_val) or pd.isna(maxSTD_val) or pd.isna(meanCOV_val) or pd.isna(maxCOV_val):
        raise ValueError("Summary values contain NaN after parsing the last two rows.")

    meanSTD = float(meanSTD_val)
    meanCOV = float(meanCOV_val)
    maxSTD = float(maxSTD_val)
    maxCoV = float(maxCOV_val)

    return meanSTD, meanCOV, maxSTD, maxCoV


def create_pointwise_std_summary(input_folder: str | os.PathLike, output_folder: Optional[str | os.PathLike] = None, recursive: bool = False) -> List[Path]:
    """Scan a folder for Excel files and build per-force summary workbooks.

    Returns a list of created summary file Paths.
    """
    input_folder = Path(input_folder)
    out_base = Path(output_folder) if output_folder else input_folder
    out_base.mkdir(parents=True, exist_ok=True)
    if not input_folder.exists():
        raise FileNotFoundError(f"Folder does not exist: {input_folder}")

    # Collect files and group by force token
    groups: Dict[str, List[Path]] = {}
    for f in _iter_excel_files(input_folder, recursive=recursive):
        name = f.name
        # Only process files that match the intended pattern segment
        if "trial_STD_CoV" not in name:
            continue
        force = _extract_force_from_filename(name)
        if not force:
            force = "unspecified"  # fallback bucket if no force token is found
        groups.setdefault(force, []).append(f)

    if not groups:
        raise RuntimeError("No matching Excel files found (expected filenames containing 'trial_STD_CoV').")

    created: List[Path] = []

    for force, files in groups.items():
        # For each sheet name, accumulate rows from all trials
        per_sheet_rows: Dict[str, List[Dict[str, object]]] = {}
        sheet_order: List[str] = []  # preserve a stable order (from first file)

        for f in sorted(files):
            trial = _extract_trial_from_filename(f.name)
            # Read all sheets in the workbook
            try:
                sheets = pd.read_excel(f, sheet_name=None, engine='openpyxl')
            except Exception:
                # Fallback for .xls or engine issues
                sheets = pd.read_excel(f, sheet_name=None)

            if not sheet_order:
                sheet_order = list(sheets.keys())

            for sheet_name, df in sheets.items():
                try:
                    meanSTD, meanCOV, maxSTD, maxCoV = _pick_summary_values(df)
                except Exception as e:
                    raise RuntimeError(f"Error processing file '{f.name}' sheet '{sheet_name}': {e}")

                row = {
                    "Trial": trial,
                    "meanSTD": meanSTD,
                    "meanCOV": meanCOV,
                    "maxSTD": maxSTD,
                    "maxCoV": maxCoV,
                }
                per_sheet_rows.setdefault(sheet_name, []).append(row)

        # Build the output path
        out_name = f"pointwise_STD_{force}_summary.xlsx"
        out_path = out_base / out_name

        # Write two sheets (one per side). If more than two exist, write all.
        with pd.ExcelWriter(out_path, engine='openpyxl') as writer:
            for sheet_name in sheet_order:
                rows = per_sheet_rows.get(sheet_name, [])
                if not rows:
                    continue
                df_out = pd.DataFrame(rows, columns=["Trial", "meanSTD", "meanCOV", "maxSTD", "maxCoV"])
                # Sort trials naturally if numeric
                try:
                    df_out["_k"] = pd.to_numeric(df_out["Trial"], errors='coerce')
                    df_out.sort_values(["_k", "Trial"], inplace=True)
                    df_out.drop(columns=["_k"], inplace=True)
                except Exception:
                    pass

                # Round all numeric values to 3 decimals (data-level), then convert CoV to fractions
                for col in ["meanSTD", "meanCOV", "maxSTD", "maxCoV"]:
                    df_out[col] = pd.to_numeric(df_out[col], errors="coerce").round(3)
                # Convert CoV columns from percentage points to fractions so that Excel's % format displays correctly
                df_out["meanCOV"] = df_out["meanCOV"] / 100.0
                df_out["maxCoV"] = df_out["maxCoV"] / 100.0

                # Write to Excel first
                df_out.to_excel(writer, index=False, sheet_name=sheet_name)

                # Access the worksheet to format as an Excel Table and set number formats
                ws = writer.sheets[sheet_name]

                from openpyxl.worksheet.table import Table, TableStyleInfo
                from openpyxl.utils import get_column_letter

                # Determine the table range
                max_row = ws.max_row
                max_col = ws.max_column
                last_col_letter = get_column_letter(max_col)
                table_ref = f"A1:{last_col_letter}{max_row}"
                tbl = Table(displayName=f"Tbl_{sheet_name}", ref=table_ref)
                style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                                       showLastColumn=False, showRowStripes=True, showColumnStripes=False)
                tbl.tableStyleInfo = style
                ws.add_table(tbl)

                # Apply number formats: show integers for STD columns, percentage for CoV columns
                # (data already rounded to 3 decimals above, but display as integers / percents)
                headers = [cell.value for cell in ws[1]]
                header_to_col = {h: i+1 for i, h in enumerate(headers)}

                # Columns to format
                int_cols = ["meanSTD", "maxSTD"]
                pct_cols = ["meanCOV", "maxCoV"]

                for hdr in int_cols:
                    if hdr in header_to_col:
                        col_idx = header_to_col[hdr]
                        for r in range(2, max_row + 1):
                            ws.cell(row=r, column=col_idx).number_format = "0"  # show full integers

                for hdr in pct_cols:
                    if hdr in header_to_col:
                        col_idx = header_to_col[hdr]
                        for r in range(2, max_row + 1):
                            ws.cell(row=r, column=col_idx).number_format = "0%"  # show integer percent

        created.append(out_path)

    return created


if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(description="Create pointwise STD/CoV summary files from Excel workbooks.")
    parser.add_argument(
        "folders",
        type=str,
        nargs="*",
        default=["/Users/noah/LRZ Sync+Share/MA/ZHS_LabView_Messungen/Exploration_V2/Shoes_and_footholds/sorted_by_shoes/front/summary"],
        help="One or more folders containing per-trial Excel files"
    )
    parser.add_argument("-r", "--recursive", action="store_true", help="Recurse into subfolders when searching for Excel files")
    parser.add_argument("-o", "--out", type=str, default=None, help="Output folder for the summary files (defaults to the input folder)")
    args = parser.parse_args()

    all_results = []
    for folder in args.folders:
        results = create_pointwise_std_summary(folder, output_folder=args.out, recursive=args.recursive)
        all_results.extend(results)

    print("Created summary files:")
    for p in all_results:
        print(" -", p)
