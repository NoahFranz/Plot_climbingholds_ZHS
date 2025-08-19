import json
import pandas as pd
from pathlib import Path
import glob
import os
import config
from copy import copy
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import PatternFill, Font
from openpyxl.formatting.rule import CellIsRule

def build_hold_tables(data, force_list, characteristics, holds):
    """
    Constructs tables of force metrics for each hold and characteristic.
    Iterates over holds and characteristics to extract relevant data from the JSON structure,
    organizing it into pandas DataFrames for later export.
    """
    hold_tables = {hold: [] for hold in holds}
    for hold in holds:
        for characteristic in characteristics:
            # If characteristic is 'mean', include 'Contacttime' in force list
            current_force_list = force_list + ['Contacttime'] if characteristic == 'mean' else force_list
            all_data = []
            for filename, data_entry in data.items():
                # Map filename to a shorter acronym if available, else use original filename
                short_filename = next(
                    (val for key, val in config.file_acronyms_map.items() if key in filename),
                    filename
                )
                row = {'filename': short_filename}
                # Skip if hold data is not present in this entry
                if hold not in data_entry:
                    continue
                intervals = data_entry[hold].get('intervals', {})
                mean_metrics = intervals.get('Mean-Metrics', {})
                for force_type in current_force_list:
                    if force_type in mean_metrics:
                        max_val = mean_metrics[force_type].get("max")
                        min_val = mean_metrics[force_type].get("min")
                        if characteristic == "max":
                            if max_val is not None and min_val is not None:
                                value = max_val if abs(max_val) >= abs(min_val) else min_val
                            elif max_val is not None:
                                value = max_val
                            elif min_val is not None:
                                value = min_val
                            else:
                                continue
                        elif characteristic == "min":
                            if max_val is not None and min_val is not None:
                                value = min_val if abs(min_val) >= abs(max_val) else max_val
                            elif min_val is not None:
                                value = min_val
                            elif max_val is not None:
                                value = max_val
                            else:
                                continue
                        else:
                            value = mean_metrics[force_type].get(characteristic)
                        if value is not None:
                            col_name = f"{force_type}_{characteristic}"
                            row[col_name] = value
                all_data.append(row)
            # Convert collected rows into a DataFrame
            df = pd.DataFrame(all_data)
            hold_tables[hold].append((characteristic, df))
    return hold_tables

# Updated function: build_interval_statistics
def build_interval_statistics(data, force_list, characteristics, holds):
    from statistics import mean, stdev
    interval_stats = {hold: {char: {} for char in characteristics} for hold in holds}
    for hold in holds:
        for filename, data_entry in data.items():
            short_filename = next(
                (val for key, val in config.file_acronyms_map.items() if key in filename),
                filename
            )
            if hold not in data_entry:
                continue
            intervals = data_entry[hold].get('intervals', {})
            for char in characteristics:
                result_row = {'file': short_filename}
                force_values = {f: [] for f in force_list}
                force_columns = {}
                for force in force_list:
                    force_columns[force] = []
                    for interval, metrics in intervals.items():
                        if interval == "Mean-Metrics":
                            continue
                        val = metrics.get(force, {}).get(char)
                        if val is not None:
                            colname = f"{force}_{char}_{interval}"
                            force_columns[force].append((colname, val))
                            force_values[force].append(val)
                for force in force_list:
                    for colname, val in force_columns[force]:
                        result_row[colname] = val
                    vals = force_values[force]
                    if vals:
                        m = mean(vals)
                        s = stdev(vals) if len(vals) > 1 else 0
                        cov = (s / m) * 100 if m != 0 else 0
                        result_row[f"{force}_{char}_mean"] = round(m)
                        result_row[f"{force}_{char}_std"] = round(s)
                        result_row[f"{force}_{char}_cov"] = f"{round(cov)}%"
                        result_row[f"{force}_{char}_"] = ""  # empty column after CoV
                interval_stats[hold][char][short_filename] = result_row
    return interval_stats

def write_table(writer, worksheet, df, characteristic, current_row):
    """
    Writes a DataFrame to an Excel worksheet starting at a specified row.
    Adds a header with the characteristic name, formats the table with styles,
    adjusts column widths, and applies number formatting to cells.
    """
    # Write DataFrame to the worksheet starting one row below current_row
    df.to_excel(writer, sheet_name=worksheet.title, index=False, startrow=current_row + 1)
    # Write the characteristic name above the table for clarity
    cell = worksheet.cell(row=current_row + 1, column=1)
    cell.value = characteristic
    # Make the characteristic header bold
    bold_font = copy(cell.font)
    bold_font.bold = True
    cell.font = bold_font

    # Define the Excel range that the table will occupy
    end_column = get_column_letter(df.shape[1])
    end_row = current_row + len(df) + 2
    table_range = f"A{current_row + 2}:{end_column}{end_row}"
    # Create an Excel table with a specific style for better readability
    table = Table(displayName=f"{worksheet.title}_{characteristic}", ref=table_range)
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                           showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    table.tableStyleInfo = style
    worksheet.add_table(table)

    # Adjust column widths based on the maximum length of the content in each column
    for i, column_cells in enumerate(
            worksheet.iter_cols(min_row=current_row + 2, max_row=end_row, min_col=1, max_col=df.shape[1]), start=1):
        max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
        worksheet.column_dimensions[get_column_letter(i)].width = max_length + 2

    # Apply number formatting to numeric cells, with special formatting for 'Contacttime_mean'
    for j, col in enumerate(worksheet.iter_cols(min_row=current_row + 2, max_row=end_row, min_col=1, max_col=df.shape[1]), start=1):
        header = worksheet.cell(row=current_row + 2, column=j).value
        for cell in col:
            if isinstance(cell.value, (int, float)):
                if header == "Contacttime_mean":
                    cell.number_format = '0.0'  # One decimal place for contact time
                else:
                    cell.number_format = '0'    # Integer format for other values
    # Return the next row index after the table plus some spacing
    return current_row + len(df) + 5

# Updated function: write_interval_statistics
def write_interval_statistics(writer, hold, interval_stats):
    worksheet = writer.book.create_sheet(title=f"{hold}_intervals")
    current_row = 0
    for char, stats in interval_stats[hold].items():
        df = pd.DataFrame(list(stats.values()))
        worksheet.cell(row=current_row + 1, column=1, value=char)
        df.to_excel(writer, sheet_name=worksheet.title, index=False, startrow=current_row + 1)
        current_row += len(df) + 4

def extract_force_data(json_path, output_excel):
    """
    Loads force data from a JSON file, builds tables for each hold and characteristic,
    and writes these tables into an Excel workbook with formatting.
    """
    import openpyxl
    with open(json_path, 'r') as f:
        data = json.load(f)

    # Define the list of force types and characteristics to extract
    force_list = ['Fy', 'Fz', 'Fx', 'Fres_xyz', 'FgR_1', 'FgR_2', 'FgR_sum','Fy_sum', 'Fz_sum', 'Fx_sum', 'Fres_xyz_sum']
    characteristics = ['max', 'mean', 'impuls', 'min', 'maxROFD']
    holds = ['G1R', 'G2L']

    # Build the tables from the JSON data
    hold_tables = build_hold_tables(data, force_list, characteristics, holds)
    interval_stats = build_interval_statistics(data, force_list, characteristics, holds)

    # Write the tables to an Excel file using openpyxl engine
    with pd.ExcelWriter(output_excel, engine='openpyxl') as writer:
        for hold, tables in hold_tables.items():
            current_row = 0
            for characteristic, df in tables:
                if not df.empty:
                    worksheet = writer.book.create_sheet(title=hold) if hold not in writer.sheets else writer.sheets[hold]
                    current_row = write_table(writer, worksheet, df, characteristic, current_row)
                else:
                    current_row += 2
            write_interval_statistics(writer, hold, interval_stats)

if __name__ == "__main__":
    import re
    # List of directories containing JSON summary files to process
    json_dirs = [
        # "/Users/noah/LRZ Sync+Share/MA/ZHS_LabView_Messungen/Exploration_V2/Shoes_and_footholds/best-grey/cross",
        # "/Users/noah/LRZ Sync+Share/MA/ZHS_LabView_Messungen/Exploration_V2/Shoes_and_footholds/best-grey/cross+front-combined",
        # "/Users/noah/LRZ Sync+Share/MA/ZHS_LabView_Messungen/Exploration_V2/Shoes_and_footholds/best-grey/front"
        #"/Users/noah/LRZ Sync+Share/MA/ZHS_LabView_Messungen/Exploration_V2/Technik/Rightside_data"
        #"/Users/noah/LRZ Sync+Share/MA/ZHS_LabView_Messungen/Exploration_V2/ForceDevRatio"
        #"/Users/noah/LRZ Sync+Share/MA/ZHS_LabView_Messungen/Exploration_V2/Clipping",
       # "/Users/noah/LRZ Sync+Share/MA/ZHS_LabView_Messungen/Exploration_V2/shakeout",
 #       "/Users/noah/LRZ Sync+Share/MA/ZHS_LabView_Messungen/Exploration_V2/shakeout/GL_2vs1FH",
 #"/Users/noah/LRZ Sync+Share/MA/ZHS_LabView_Messungen/Exploration_V2/shakeout/GR_2vs1FH",
 #"/Users/noah/LRZ Sync+Share/MA/ZHS_LabView_Messungen/Exploration_V2/shakeout/Switch"
"/Users/noah/LRZ Sync+Share/MA/ZHS_LabView_Messungen/Exploration_V2/Shoes_and_footholds/best-grey/front",
"/Users/noah/LRZ Sync+Share/MA/ZHS_LabView_Messungen/Exploration_V2/Shoes_and_footholds/worst-black/front",
"/Users/noah/LRZ Sync+Share/MA/ZHS_LabView_Messungen/Exploration_V2/Shoes_and_footholds/medium-yellow/front",
"/Users/noah/LRZ Sync+Share/MA/ZHS_LabView_Messungen/Exploration_V2/Shoes_and_footholds/sorted_by_shoes/front/Trail",
"/Users/noah/LRZ Sync+Share/MA/ZHS_LabView_Messungen/Exploration_V2/Shoes_and_footholds/sorted_by_shoes/front/Perf",
"/Users/noah/LRZ Sync+Share/MA/ZHS_LabView_Messungen/Exploration_V2/Shoes_and_footholds/sorted_by_shoes/front/HighEnd",
"/Users/noah/LRZ Sync+Share/MA/ZHS_LabView_Messungen/Exploration_V2/Shoes_and_footholds/sorted_by_shoes/front/Basic",
 # "/Users/noah/LRZ Sync+Share/MA/ZHS_LabView_Messungen/Exploration_V2/Moment"
        # "/Users/noah/LRZ Sync+Share/MA/ZHS_LabView_Messungen/Exploration_V2/Reliability/071_endurace"
   
    ]

    # Discover all JSON summary files in the specified directories
    json_paths = []
    for dir_path in json_dirs:
        json_paths.extend(glob.glob(os.path.join(dir_path, "*summary.json")))
    json_paths = list(set(json_paths))  # Optional: remove duplicates

    # Process each JSON file and export the extracted data to an Excel file
    for json_path in json_paths:
        json_file = Path(json_path)
        match = re.match(r'^(\d+)_\D', json_file.stem + '_')
        file_number = match.group(1) if match else ''
        output_excel = json_file.with_name(json_file.stem + '_Mean-Metrics.xlsx')
        # Extract force data and write to Excel
        extract_force_data(json_path, output_excel)
        print(f"Exported {output_excel}")